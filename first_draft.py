import pandas as pd
import openpyxl

# Load the Excel file
filepath = 'client.xlsx'
df = pd.read_excel(filepath)

# Clean and normalize the 'PR Award Number' column (remove leading/trailing spaces)
df['PR Award Number'] = df['PR Award Number'].str.strip()

# Exclude columns from highlighting
cols_to_exclude = ['Screener Initials', 'Other Comments']

# Sort by PR Award Number in ascending order
df.sort_values(by='PR Award Number', inplace=True)

# Create a new Excel workbook to save the flagged results
wb = openpyxl.load_workbook(filepath)
ws = wb.active

# Function to scan for discrepancies and add "FLAG" row
def scan_and_flag(group):
    columns = group.columns
    flag_added = False
    for col in columns:
        if col == 'PR Award Number' or col in cols_to_exclude:
            continue
        values = group[col]
        first_value = values.iloc[0]
        for idx in range(1, len(group)):
            current_value = values.iloc[idx]
            if current_value != first_value:
                if not flag_added:
                    # Insert a new row under the cell with the first discrepancy and mark it with "FLAG"
                    row_idx = group.index[0]  # First row of the group
                    ws.insert_rows(row_idx + 1)  # +1 to insert under the row with the first discrepancy
                    flag_cell = ws.cell(row=row_idx + 2, column=columns.get_loc(col) + 1)  # +2 for newly inserted row and +1 for 1-based index
                    flag_cell.value = "FLAG"
                    flag_added = True
                # Clear the current row after marking "FLAG" for the group
                cell = ws.cell(row=group.index[idx] + 1, column=columns.get_loc(col) + 1)  # +1 for 1-based index
                cell.value = ""

# Group by PR Award Number
grouped = df.groupby('PR Award Number')

# Apply the function to each group
for name, group in grouped:
    scan_and_flag(group)

# Save the workbook with the flagged discrepancies
output_file_path = 'flagged_with_flags.xlsx'
wb.save(output_file_path)

print("Discrepancies flagged and saved to", output_file_path)
