import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

filepath = 'client.xlsx'
df = pd.read_excel(filepath)
wb = openpyxl.load_workbook(filepath)
ws = wb.active
grouped = df.groupby('PR Award Number')
#change later
colors = [
    "FFFF0000", "FF00FF00", "FF0000FF", "FFFFFF00", "FFFF00FF", "FF00FFFF",
    "FF800000", "FF808000", "FF008000", "FF800080", "FF008080", "FF000080",
    "FFFFA500", "FF7FFFD4", "FFDA70D6", "FF00CED1", "FF00FA9A", "FFBA55D3",
    "FFA52A2A", "FFDEB887", "FF5F9EA0", "FF7FFF00", "FFD2691E", "FF6495ED",
    "FFDC143C", "FF00FFFF", "FF00008B", "FF008B8B", "FFB8860B", "FFA9A9A9",
    "FF006400", "FFBDB76B", "FF8B008B", "FF556B2F", "FFFF8C00", "FF9932CC",
    "FF8B0000", "FFE9967A", "FF8FBC8F", "FF483D8B", "FF2F4F4F", "FFFF1493",
    "FF00BFFF", "FF696969", "FF696969", "FF1E90FF", "FFB22222", "FF228B22"
]

fills = [PatternFill(start_color=color, end_color=color, fill_type="solid") for color in colors]

color_map = {}

exclude_columns = ['Screener Initials', 'Other Comments']

#func
def highlight_discrepancies(group, color_fill):
    columns = group.columns
    for col in columns:
        if col == 'PR Award Number' or col in exclude_columns:
            continue
        values = group[col]
        first_value = values.iloc[0]
        if any(values != first_value):
            for idx in group.index:
                current_value = values.loc[idx]
                if pd.isna(current_value) or str(current_value).lower() == 'comments': #to ignore
                    continue
                cell = ws.cell(row=idx+2, column=columns.get_loc(col)+1) 
                cell.fill = color_fill
color_idx = 0
for name, group in grouped:
    color_fill = fills[color_idx % len(fills)]
    color_map[name] = colors[color_idx % len(colors)]
    highlight_discrepancies(group, color_fill)
    color_idx += 1

start_row = ws.max_row + 2
ws.cell(row=start_row, column=1).value = "PR Award Number"
ws.cell(row=start_row, column=2).value = "Color Code"
for i, (award_number, color) in enumerate(color_map.items()):
    ws.cell(row=start_row + i + 1, column=1).value = award_number
    ws.cell(row=start_row + i + 1, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
output_file_path = 'flagged_colorcoded.xlsx'
wb.save(output_file_path)

print("Discrepancies highlighted and saved to", output_file_path)
